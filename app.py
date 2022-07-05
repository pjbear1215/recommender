from openpyxl import load_workbook
wb = load_workbook(filename = 'sample.xlsx')
sheet_ranges = wb['Sheet1']
print(sheet_ranges['A1'].value)
#from openpyxl import Workbook
#wb = Workbook()

# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample.xlsx")
